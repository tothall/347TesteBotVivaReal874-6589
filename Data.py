import time

from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from time import sleep
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, NoSuchElementException
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook


class MetaData:

    @staticmethod
    def get_extraction_info(wait, driver, time_start, time_stop, input_last_page, current_last_page, extracted_ads):
        wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='results-summary__wrapper js-wrapper']")))
        property_type = driver.find_element(By.XPATH, "//li[@class='hbs-filters__pill js-filter']").text
        ads_quantity = driver.find_element(By.XPATH, "//strong[@class='results-summary__count js-total-records']").text
        extracted_pages = f"{current_last_page} Páginas extraídas de {input_last_page}"
        extracted_ads = f"{extracted_ads} Anúncios extraídos de {ads_quantity}"

        date = datetime.date.today()
        date_time_hour = datetime.datetime.now().hour
        date_time_minute = datetime.datetime.now().minute
        date_time_second = datetime.datetime.now().second
        file_datetime = f"{date} - {date_time_hour}h {date_time_minute}m {date_time_second}s"
        elapsed_time = time_start - time_stop

        metadata = [date, time_start, time_stop, elapsed_time, property_type, extracted_ads, extracted_pages]

        wb = Workbook()
        ws = wb.active
        ws.append(["Informações da Extração"])
        ws.append(["Data", "Hora Início", "Hora Término", "Tempo Gasto", "Tipo de Imóvel",
                   "Quantidade de Anúncios Extraídos", "Quantidade de Páginas Extraídas"])
        ws.append([metadata])
        ws.append([""])
        # Cabeçalho atributos dos imóveis
        ws.append(["Endereço do Imóvel", "Área em m2", "Quantidade de Quartos", "Quantidade de Banheiros",
                   "Quantidade de Vagas", "Descrição do Imóvel", "Preço do Imóvel(R$)", "Link Anúncio"])

        ads_data = load_workbook(filename='Data_Extraction_Temp.xlsx')
        for i in range(len(ads_data)):
            ws.append(ads_data[i])
        wb.save(f"Extração Viva Real {file_datetime}.xlsx")


class Extraction:
    """Todas as funções dessa classe recebem como argumento o wait e/ou o argumento driver.
        Isso acontece pelo fato das funções dependerem do objeto driver e do objeto wait da classe
        Configurations"""
    @staticmethod
    def all_page_houses_extraction(wait, driver):

        house_addresses = []
        house_areas = []
        house_rooms = []
        house_bathrooms = []
        house_garages = []
        house_descriptions = []
        house_prices = []
        house_urls = []
        houses_data = []
        ads_quantity = 0

        # Localizar todos os imóveis da página
        houses_div = driver.find_element(By.XPATH, "//section[@class='results__main']")
        wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='property-card__content']")))
        houses_data = houses_div.find_elements(By.XPATH, "//div[@class='property-card__content']")

        # Extração dos atributos de cada imóvel
        for i in range(len(houses_data)):
            try:
                # Obter a lista atualizada de elementos após cada clique
                houses_data = houses_div.find_elements(By.XPATH, "//div[@class='property-card__content']")

                # Verificar se o índice ainda é válido
                if i >= len(houses_data):
                    break  # Todos os imóveis foram clicados/extraídos

                '''# Obter o link do imóvel usando índice
                a = houses_data[i].find_element(By.XPATH, "//a[@class='property-card__content-link js-card-title']")
                product_url = a.get_attribute("href")

                # Navegar para a página do produto usando o URL
                driver.get(product_url)
                Randomization.random_page_wait()
                # Voltar para a página anterior para evitar problemas com elementos "stale"
                #Randomization.random_time_actions()
                driver.execute_script("window.history.go(-1);")'''

                # Extraindo os atributos de cada imóvel

                soup = BeautifulSoup(driver.page_source, "lxml")

                # Endereço do imóvel

                house_address_elements = soup.find_all("span", class_="property-card__address")
                for element in house_address_elements:
                    house_addresses.append(element.get_text())

                # Metragem dos imóveis

                house_area_elements = soup.find_all("li", class_="property-card__detail-item property-card__detail-area")
                for element in house_area_elements:
                    house_areas.append(element.get_text())

                # Quantidade de quartos

                house_rooms_elements = soup.find_all("li", class_="property-card__detail-item property-card__detail-room js-property-detail-rooms")
                for element in house_rooms_elements:
                    house_rooms.append(element.get_text())

                # Quantidade de banheiros

                house_bathrooms_elements = soup.find_all("li", class_="property-card__detail-item property-card__detail-bathroom js-property-detail-bathroom")
                for element in house_bathrooms_elements:
                    house_bathrooms.append(element.get_text())

                # Quantidade de vagas

                house_garage_elements = soup.find_all("li", class_="property-card__detail-item property-card__detail-garage js-property-detail-garages")
                for element in house_garage_elements:
                    house_garages.append(element.get_text())

                # Descrição do imóvel

                house_description_elements = soup.find_all("span", class_="property-card__title js-cardLink js-card-title")
                for element in house_description_elements:
                    house_descriptions.append(element.get_text().strip())

                # Preço do imóvel

                house_price_elements = soup.find_all("div", class_="property-card__price js-property-card-prices js-property-card__price-small")
                for element in house_price_elements:
                    house_prices.append(element.get_text())

                # Link de acesso ao anúncio do imóvel

                a = houses_data[i].find_element(By.XPATH, "//a[@class='property-card__content-link js-card-title']")
                house_url = a.get_attribute("href")
                house_urls.append(house_url)

                # Armazenando os atributos do imóvel na lista houses_data
                houses_data = list(zip(house_addresses, house_areas, house_rooms, house_bathrooms, house_garages, house_descriptions, house_prices, house_urls))
                # Salva os atributos de cada imóvel na planilha
                wb = Workbook()
                ws = wb.active
                for j in range(len(houses_data)):
                    ws.append(houses_data[i])
                wb.save("Data_Extraction_Temp.xlsx")

                ads_quantity += 1

                # Esperar até que os elementos estejam presentes novamente
                wait.until(EC.presence_of_all_elements_located((By.XPATH, "//section[@class='results__main']")))
                wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='property-card__content']")))
                sleep(2)  # Aguardar um curto período para evitar problemas de renderização
            except (TimeoutException, StaleElementReferenceException):
                print("Elemento tornou-se stale ou a espera excedeu o limite. Tentando próximo elemento.")
                continue

        return ads_quantity
