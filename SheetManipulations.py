import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from Data import MetaData


class SheetManipulations:

    @staticmethod
    def extraction_headers():
        wb = Workbook()
        ws = wb.active
        # Cabeçalho metadados da extração
        date_time = datetime.date.today()
        date_time_hour = datetime.datetime.now().hour
        date_time_minute = datetime.datetime.now().minute
        date_time_second = datetime.datetime.now().second
        file_datetime = f"{date_time} - {date_time_hour}h {date_time_minute}m {date_time_second}s"
        ws.append(["Informações da Extração"])
        ws.append(["Data", "Hora Início", "Hora Término", "Tempo Gasto", "Tipo de Imóvel", "Quantidade de Anúncios Encontrados", "Intervalo de Páginas Raspadas"])
        ws.append(MetaData.get_extraction_info())
        ws.append([""])
        # Cabeçalho atributos dos imóveis
        ws.append(["Endereço do Imóvel", "Área em m2", "Quantidade de Quartos", "Quantidade de Banheiros", "Quantidade de Vagas", "Descrição do Imóvel", "Preço do Imóvel(R$)", "Link Anúncio"])
        wb.save(f"Extração Viva Real {file_datetime}.xlsx")
        file_name = f"Extração Viva Real {file_datetime}.xlsx"
        return file_name

    @staticmethod
    def save_to_xlsx(data):
        """Armazena os dados dos imóveis numa planilha"""
        wb = Workbook()
        ws = wb.active
        for i in range(len(data)):
            ws.append(data[i])
        wb.save("Anuncios_Casas_Viva_Real.xlsx")

print(SheetManipulations.extraction_headers())
